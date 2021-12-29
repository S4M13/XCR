import openpyxl as xl
import shutil
import datetime

if __name__ == "__main__":
    from ..app import Settings
    import Database
    from GlobalContext import GlobalContext
    import Database
else:
    import Settings
    from Util import Database
    from Util.GlobalContext import GlobalContext
    from Util import Database


def generate_overall_analysis():
    """
    This function carries out analysis on the database and creates a spreadsheet in Exports which is to
    be downloaded by users in order to carry out more in-depth analysis than is available in the
    main menus

    :return: The name of the generated file
    """

    # Generate the new name
    time_stamp = datetime.datetime.now().strftime("%b %e %H-%M")
    file_name = "Overall Analysis " + time_stamp + ".xlsx"
    location = Settings.EXPORTS / Settings.CURRENT_OVERALL

    # Copy the base template across
    shutil.copyfile(Settings.SPREADSHEETS / Settings.OVERALL_ANALYSIS_SPREADSHEET, location)

    # Work out what the headers are
    headers = GlobalContext.STUDENTS_DATASTORE.DataHeaders.headers

    for club in GlobalContext.CLUBS_DATASTORE.data:
        headers.append(club[1] + " Attendance")  # TODO : Check UI/Name formatting

    # Load the XL workbook
    wb = xl.load_workbook(filename=location, data_only=True)
    sheet = wb["Analysis"]

    # Add the headers
    for index, entry in enumerate(headers):
        sheet.cell(row=1, column=index+1).value = entry

    # Pre-Calculate the data about attendances

    records = Database.Record.query.all()  # Fetch all records
    analysis = {}

    for record in records:
        if record.student_uid in analysis.keys():
            if record.club_uid in analysis[record.student_uid]:
                analysis[record.student_uid][record.club_uid] += 1
            else:
                analysis[record.student_uid][record.club_uid] = 1
        else:
            analysis[record.student_uid] = {record.club_uid: 1}

    # Iterate over the students, adding their calculated entries

    for index, student_entry in enumerate(GlobalContext.STUDENTS_DATASTORE.data):
        row = index + 2
        column = 1

        # Add the data which already exists about the student
        for data in student_entry:
            sheet.cell(row=row, column=column).value = data
            column += 1

        # Iterate over the clubs and calculate and add the data
        for club_entry in GlobalContext.CLUBS_DATASTORE.data:
            result = 0

            if student_entry[0] in analysis.keys():
                data_point = analysis[student_entry[0]]
                if club_entry[0] in data_point.keys():
                    result = data_point[club_entry[0]]

            sheet.cell(row=row, column=column).value = result
            column += 1

    wb.save(location)
    return file_name


def generate_student_analysis(student_uid):
    """
    This function carries out analysis on the database and exports the data as XL spreadsheet to be downloaded
    by the user to carry out further analysis which cannot be done from the main page

    :param student_uid: The UID of the student to analyse. The UID must be checked to ensure it is valid.
    :return: The name of the generated file [or False if the UID is invalid, or returns multiple students]
    """

    # Get the students information
    student_info = GlobalContext.STUDENTS_DATASTORE.return_specific_entries("UID", student_uid)

    if len(student_info) != 1:
        return False

    student_info = student_info[0]

    # Generate the new name
    time_stamp = datetime.datetime.now().strftime("%b %e %H-%M")
    file_name = f"{student_info[1]} {student_info[2]} Analysis " + time_stamp + ".xlsx"
    location = Settings.EXPORTS / Settings.CURRENT_STUDENT

    # Copy the base template across
    shutil.copyfile(Settings.SPREADSHEETS / Settings.OVERALL_ANALYSIS_SPREADSHEET, location)

    # Load the XL workbook
    wb = xl.load_workbook(filename=location, data_only=True)
    sheet = wb["Analysis"]

    # Fetch all the relevant records
    records = Database.Record.query.filter_by(student_uid=student_uid).all()

    # Work out the date of the sessions and the clubs that were attended

    sessions = []
    clubs = []

    for record in records:
        weekly_stamp = record.attendance_date.strftime("%W-%Y")

        if weekly_stamp not in sessions:
            sessions.append(weekly_stamp)

        if record.club_uid not in clubs:
            clubs.append(record.club_uid)

    sessions.sort()
    clubs.sort()

    # Work out the headers
    headers = GlobalContext.CLUBS_DATASTORE.DataHeaders.headers + sessions

    # Add the headers
    for index, entry in enumerate(headers):
        sheet.cell(row=1, column=index + 1).value = entry

    # Add clubs data
    for index, club_entry in enumerate(GlobalContext.CLUBS_DATASTORE.data):
        row = index + 2
        column = 1

        # Add the data which already exists about the club
        for data in club_entry:
            sheet.cell(row=row, column=column).value = data
            column += 1

        # Loop over records and add attendance marks
    column_offset = 1 + len(GlobalContext.CLUBS_DATASTORE.DataHeaders.headers)
    adjusted_clubs = [x[0] for x in GlobalContext.CLUBS_DATASTORE.data]

    for record in records:
        # Fetch correct column and row
        column = sessions.index(record.attendance_date.strftime("%W-%Y")) + column_offset
        row = adjusted_clubs.index(record.club_uid) + 2

        if current := sheet.cell(column=column, row=row).value == "":
            sheet.cell(column=column, row=row).value = 1
        else:
            sheet.cell(column=column, row=row).value = int(current) + 1

    wb.save(location)
    return file_name


def generate_club_analysis(club_uid):
    """
    This function carries out analysis on the database and exports the data as XL spreadsheet to be downloaded
    by the user to carry out further analysis which cannot be done from the main page

    :param club_uid: The UID of the club to analyse. The UID must be checked to ensure it is valid.
    :return: The name of the generated file [or False if the UID is invalid, or returns multiple clubs]
    """

    # Get the clubs information
    club_info = GlobalContext.CLUBS_DATASTORE.return_specific_entries("UID", club_uid)

    if len(club_info) != 1:
        return False

    club_info = club_info[0]

    # Generate the new name
    time_stamp = datetime.datetime.now().strftime("%b %e %H-%M")
    file_name = f"{club_info[1]} Analysis " + time_stamp + ".xlsx"
    location = Settings.EXPORTS / Settings.CURRENT_CLUB

    # Copy the base template across
    shutil.copyfile(Settings.SPREADSHEETS / Settings.OVERALL_ANALYSIS_SPREADSHEET, location)

    # Load the XL workbook
    wb = xl.load_workbook(filename=location, data_only=True)
    sheet = wb["Analysis"]

    # Fetch all the relevant records
    records = Database.Record.query.filter_by(club_uid=club_uid).all()

    # Work out the date of the sessions and the students that attended

    sessions = []
    students = []

    for record in records:
        weekly_stamp = record.attendance_date.strftime("%W-%Y")

        if weekly_stamp not in sessions:
            sessions.append(weekly_stamp)

        if record.student_uid not in students:
            students.append(record.student_uid)

    sessions.sort()
    students.sort()

    # Work out the headers
    headers = GlobalContext.STUDENTS_DATASTORE.DataHeaders.headers + sessions

    # Add the headers
    for index, entry in enumerate(headers):
        sheet.cell(row=1, column=index + 1).value = entry

    # Add student data
    for index, student_entry in enumerate(GlobalContext.STUDENTS_DATASTORE.data):
        row = index + 2
        column = 1

        # Add the data which already exists about the student
        for data in student_entry:
            sheet.cell(row=row, column=column).value = data
            column += 1

    # Loop over records and add attendance marks
    column_offset = 1 + len(GlobalContext.STUDENTS_DATASTORE.DataHeaders.headers)
    adjusted_students = [x[0] for x in GlobalContext.STUDENTS_DATASTORE.data]

    for record in records:
        # Fetch correct column and row
        column = sessions.index(record.attendance_date.strftime("%W-%Y")) + column_offset
        row = adjusted_students.index(record.student_uid) + 2

        if current := sheet.cell(column=column, row=row).value == "":
            sheet.cell(column=column, row=row).value = 1
        else:
            sheet.cell(column=column, row=row).value = int(current) + 1

    wb.save(location)
    return file_name
