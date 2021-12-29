import openpyxl as xl
from typing import Callable
import os


class DatastoreHeaders:

    def __init__(self, data: object) -> None:
        self._headers = []

        self._load_from_wb(data)

    def __str__(self) -> str:
        return str(self._headers)

    def __getitem__(self, item: int) -> str:
        return self._headers[item]

    def __setitem__(self, key: int, value: object) -> None:
        self._headers[key] = value

    def __eq__(self, other: object) -> bool:
        if type(other) != DatastoreHeaders: return False
        for header in self._headers:
            if header not in other.headers:
                return False
        return True

    def __contains__(self, item: object) -> bool:
        return self._headers.__contains__(item)

    def index_of(self, strata: str) -> int:
        """
        Returns the index of the given strata in the current datastore object.

        :param strata: The strata
        :return: The index
        """
        return self._headers.index(strata)

    def _load_from_wb(self, data: object) -> None:
        """
        Attempts to load headers from the given Workbook

        :param data: The sheet to load the headers from
        :raises Exception: Will raise an exception if headers have already been loaded for this instance
        :raises Exception: Will raise an exception if no headers can be found in the given data sheet
        """
        if len(self._headers) != 0: raise Exception("Headers have already been loaded for this instance.")

        self._headers = DatastoreHeaders._get_headers(data)

        if len(self._headers) == 0: raise Exception("Failed to find any headers in the specific data sheet.")

    @property
    def headers(self) -> list:
        return self._headers

    @staticmethod
    def _get_headers(data: object) -> list:
        """
        Returns a list of headers from the given sheet.

        :param data: The sheet to load the headers from
        :return: A list of headers
        """
        headers = []
        current_index = 1

        while True:
            val = data.cell(row=1, column=current_index).value
            if val is not None:
                headers.append(val)
            else:
                break

            current_index += 1

        return headers


class Datastore:

    def __init__(self):
        self.DataHeaders = None
        self.loaded_sheets = []
        self.loaded_sheet_names = []

        self.data = []

    def _valid_sheet(self, workbook: object) -> bool:
        """
        Checks if a workbook has a data sheet and matches the current headers.

        :param workbook: The workbook to check
        :return: A boolean representing whether the sheet is valid or not
        """
        if not ("Data" in workbook.sheetnames): return False
        data = workbook["Data"]

        if self.DataHeaders is not None:  # TODO : Add check for 'UID' in column 1
            compare_headers = DatastoreHeaders(data)

            if compare_headers != self.DataHeaders:
                return False
        return True

    def load_sheet(self, name: str, sheet: object) -> None:
        """
        Attempts to load a sheet into the datastore.

        :param name: The name fo the sheet being loaded in
        :param sheet: The sheet to load into the datastore
        :raises Exception: Raises an exception if the sheet is not valid. See _valid_sheet(sheet)
        """
        if not self._valid_sheet(sheet): ReferenceError("Failed to load an invalid sheet for this Database.")

        data = sheet["Data"]
        if self.DataHeaders is None:
            self.DataHeaders = DatastoreHeaders(data)

        self.loaded_sheets.append(sheet)
        self.loaded_sheet_names.append(name)

        current_row = 2
        while True:
            data_point = []

            if data.cell(row=current_row, column=1).value is None: break

            for index, header in enumerate(self.DataHeaders.headers):
                value = data.cell(row=current_row, column=index + 1).value
                data_point.append(value)

            self.data.append(data_point)
            current_row += 1

    def load_directory(self, location: str) -> None:
        """
        Attempts to load all the sheets from a given directory into the datastore

        :param location: The directory to load the sheets from
        :raises ReferenceError: Raises an exception if a given sheet is not valid.
        """

        files = os.listdir(location)
        for file in files:
            if file[-5:] == ".xlsx":
                wb_path = os.path.join(location, file)
                wb = xl.load_workbook(wb_path, data_only=True)
                try:
                    self.load_sheet(file, wb)
                except ReferenceError:
                    pass

    def purge_memory(self) -> None:
        """
        Removes all data in the datastore.
        """
        self.DataHeaders = None
        self.loaded_sheets = []
        self.loaded_sheet_names = []

        self.data = []

    def return_specific_entries(self, strata: str, match: str) -> list:
        """
        Searches the datastore and returns any matching entries for the given strata.

        :param strata: The strata to compare against
        :param match: The string to compare the given strata to
        :return: A list of matches
        """
        if self.DataHeaders is None: raise Exception("No headers detected. Likely, no data has been loaded.")
        if strata not in self.DataHeaders: raise Exception("Strata '{0}' is not a valid header".format(strata))

        strata_index = self.DataHeaders.index_of(strata)

        valid_data = []

        for data_point in self.data:
            if str(data_point[strata_index]) == str(match):
                valid_data.append(data_point)
        return valid_data

    def return_generic_entries(self, validator: Callable) -> list:
        """
        Searches the datastore and returns any matching entries according to the validator.

        :param validator: A function which returns a boolean representing whether the entry should be returned
        :return: A list of matches
        """
        if self.DataHeaders is None: raise Exception("No headers detected. Likely, no data has been loaded.")

        valid_data = []

        for data_point in self.data:
            if validator(data_point):
                valid_data.append(data_point)
        return valid_data
